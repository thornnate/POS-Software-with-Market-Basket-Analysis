import mysql.connector
import datetime
from tabulate import tabulate
import os
import openpyxl


# making connection to the user 
mydb = mysql.connector.connect(
    host = 'localhost',
    user = 'proj',
    password = 'proj123',
    database = 'project'
)

mc = mydb.cursor()

# create a date time instance
datetime = datetime.datetime.today()
# retrive date from date time instance  
date = datetime.date()
date = str(date)

# function to prevent empty values
def valid_input(text):
    not_valid = True
    res = ''
    while not_valid:
        res = input(text)
        res = res.strip()
        if res.split():  # if text is empty or only spaces, this creates an empty list evaluated at False
            not_valid = False
    return res

def valid_val(text):
    not_valid = True
    res = ''
    while not_valid:
        res = input(text)
        while res.isdigit() == False:
            res = input("Enter a valid number: ")
            if res.isdigit() == False:
                print("Wrong input.")
            else:
                break
        res = res.strip()
        a = res.split()
        if int(a[0]) not in range(1,):  # if input is less than or equal to 0, the function will keep on asking for an input
            not_valid = False
    return res


class employee:

    # function to take orders
    def order(self):
        bill_amt = 0
        flag = 0
        item = 0
        a = []
        b = []
        wb = openpyxl.load_workbook(r'C:\Users\LENEVO\Desktop\PROJECT\data_set.xlsx')  
        sheet = wb.active  
        size = sheet.dimensions
        row = int(size[4:]) + 1

        while(flag != str(1)):
            item += 1
            itm_id = valid_input("Enter item number or press '0' to exit: ")
            if itm_id == "0":
                item -= 1
                break
            
            # checking if item exists
            sql_fetch = "select it_no from menu where it_no = (%s)"
            mc.execute(sql_fetch,[itm_id])
            val0 = mc.fetchall()
            if val0 == []:
                print("\nItem does not exists.")
                item -= 1
                continue
            itm_qty = int(valid_val("Enter quantity: "))
            sql_fetch = "select * from menu where it_no = (%s)"
            mc.execute(sql_fetch,[itm_id])
            val = mc.fetchall()
            for i in val:
                itm_no = i[0]
                itm_name = i[1]
                itm_rate = int(i[2])
                bill_amt += (itm_rate * itm_qty)
                a.append((itm_name,itm_qty,itm_rate))
                b.append((itm_name,itm_qty,itm_rate,itm_no,date))
           
            # adding data to the excel file of data_Set   
            sql_fetch2 = "select it_name from menu"
            mc.execute(sql_fetch2)
            val2 = mc.fetchall()
            i3 = []
            for i2 in val2:
                for j in i2:
                    i3.append(j)
            if int(itm_no) == 1:
                sheet.cell(row=row,column=1).value = i3[0]
            elif int(itm_no) == 2:
                sheet.cell(row=row,column=2).value = i3[1]
            elif int(itm_no) == 3:
                sheet.cell(row=row,column=3).value = i3[2]
            elif int(itm_no) == 4:
                sheet.cell(row=row,column=4).value = i3[3]
            elif int(itm_no) == 5:
                sheet.cell(row=row,column=5).value = i3[4]    
        
            wb.save(r'C:\Users\LENEVO\Desktop\PROJECT\data_set.xlsx')
            flag = input("To add more items press 'Enter', else press '1': ")

        # writing the data in the table orders
        # if no item selected then return else add item to database
        if item == 0:
            return
        sql = "insert into orders (sum,date) values (%s,%s)"
        data = [(bill_amt,date)]
        mc.executemany(sql,data)
        
        # fetching the last order_id from database
        sql_fetch2 = "select order_id from orders order by order_id desc limit 1"
        mc.execute(sql_fetch2)
        val = mc.fetchone()
        order_no = val[0]
        order_id = tuple(str(order_no).split(" "))
        
        # inserting order details in table order_dtl
        for i in range(item):
            sql2 = "insert into order_dtl (order_id,itm_name,quantity,rate,it_no,date) values (%s,%s,%s,%s,%s,%s)"
            values = [sum((order_id,b[i]),())]
            mc.executemany(sql2,values)
        print("\nOrder placed successfully.")
        mydb.commit()

        # creating a bill in the form of a text file
        path = (r"C:\Users\LENEVO\Desktop\PROJECT\bill\ ") + str(order_no) + ".txt"
        with open(path,'w') as file_obj:
            file_obj.write("\t\tBILL NO: " + str(order_no) +  "\n\n")
            file_obj.write("DATE: " + str(date) + "\n\n")
            file_obj.write(tabulate(a, headers = ["Item_Name","Quantity","Item_Price"]) + "\n")
            file_obj.write("_____________________________________")
            file_obj.write("\n" + "Total ""\t\t\t\t  " + str(bill_amt))

    # function to edit order
    def edit_order(self):
        # deleting an item from the bill
        flag = -1
        ttl_price = 0
        bill_id = valid_val("Enter bill number to update: ")
        # check if bill exists or not
        sql_check = "select order_id from orders"
        mc.execute(sql_check)
        id = mc.fetchall()
        k = []
        for i in id:
            for j in i:
                k.append(j)

        if int(bill_id) not in k:
            print("\nBill does not exists.")
            return
        
        while(flag != str(1)):
            itm_no = valid_val("Enter item number to delete: ")
            sql = "select it_no from order_dtl where order_id = (%s)"
            mc.execute(sql,[bill_id])
            itm_data = mc.fetchall()
            c = []
            for i in itm_data:
                c.append(i[0])

            # checking if item exists in the order
            if int(itm_no) not in c:
                print("\nItem Does not exists in the order.\n")
                break
            else: 
                sql1 = "select rate, quantity from order_dtl where order_id = (%s) and it_no = (%s)"
                val = [bill_id,itm_no]
                mc.execute(sql1,val)
    
                itm_val = mc.fetchall()
                for i in itm_val:
                    rate = i[0]
                    qty = i[1]
                ttl_price += (rate * qty)    
    
                sql2 = "delete from order_dtl where order_id = (%s) and it_no = (%s)"
                value = [(bill_id,itm_no)]
                mc.executemany(sql2,value)
    
                # updating bill total
                sql3 = "select sum from orders where order_id = (%s)"
                mc.execute(sql3,[bill_id])
                bill_total = mc.fetchall()
                for i in bill_total:
                    for j in i:
                        total = j
    
                flag = input("To delete more items press 'Enter', else press '1': ")
        total -= ttl_price
        
        # updating values in the tables
        sql4 = "update orders set sum = (%s) where order_id = (%s)"
        updated_val = [(total,bill_id)]
        mc.executemany(sql4,updated_val)

        # updating the text file of the bill
        sql_fetch1 = "select * from order_dtl where order_id = (%s)"
        mc.execute(sql_fetch1,[bill_id])
        val = mc.fetchall()
        a = []
        bill_amt = 0
        for i in val:
            itm_name = i[1]
            itm_qty = int(i[2])
            itm_rate = int(i[3])
            itm_no = i[4]
            bill_amt += (itm_rate * itm_qty)
            a.append((itm_name,itm_qty,itm_rate))

        sql_fetch2 = "select sum from orders where order_id = (%s)"
        mc.execute(sql_fetch2,[bill_id])
        val = mc.fetchall()
        for i in val:
            updated_total = i[0]
            if updated_total == None:
                updated_total = 0

        sql5 = "delete from orders where sum = '0.0'"
        mc.execute(sql5)
        mydb.commit()        

        # updating the text bill 
        path = (r"C:\Users\LENEVO\Desktop\PROJECT\bill\ ") + str(bill_id) + ".txt"
        if updated_total == 0:
            os.remove(path)
        else:
            with open(path,'w') as file_obj:
                file_obj.write("\t\tBILL NO: " + str(bill_id) +  "\n\n")
                file_obj.write("DATE: " + str(date) + "\n\n")
                file_obj.write(tabulate(a, headers = ["Item_Name","Quantity","Item_Price"]) + "\n")
                file_obj.write("_____________________________________")
                file_obj.write("\n" + "Total ""\t\t\t\t  " + str(updated_total))   

    # function to print bill
    def print_bill(self):
        # fetching the last order_id from database
        sql_fetch2 = "select order_id from orders order by order_id desc limit 1"
        mc.execute(sql_fetch2)
        val = mc.fetchone()
        order_no = val[0]

        path = (r"C:\Users\LENEVO\Desktop\PROJECT\bill\ ") + str(order_no) + ".txt"
        with open(path) as file_obj:
            bill = file_obj.read()
            print()
            print(bill)
            print()
     

e1 = employee()
# e1.order()
# e1.edit_order()
# e1.print_bill()